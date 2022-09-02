"""
Прочитать сохранённый csv-файл из задания №17 и сохранить данные в excel-файл, кроме возраста – столбец с
 этими данными не нужен.

К заданию прикреплён пример как должно выглядеть содержания итогового файла.




Внимательно прочитайте задание и посмотрите прикреплённую к нему картинку. 
Данные в таблице необходимо перевернуть: данные про каждого человека необходимо
читать не слева-направо, а сверху-вниз.

"""

from random import randint

import csv
import openpyxl


def rand_id_gen():
    id_list = []
    for i in range(5):
        id_users = randint(1000,9990)
        id_list.append(id_users)
    return id_list

def person_num():
    users_pers = []
    for i in range(5):
        users = f"Person{i+1}"
        users_pers.append(users)
    return users_pers


book = openpyxl.Workbook()
sheet = book.active

user_pers = person_num()


id_list = rand_id_gen() 


def create_add_row(row_number,your_column,id_list):
    for i, value in enumerate(id_list, start=your_column):
        sheet.cell(column=i, row=row_number).value = value 


with open("data_name_number.csv",newline="") as csv_file:
    xlsx_file = csv.reader(csv_file)
    print(type(xlsx_file))
    # sheet.insert_cols(1)
    
    
    
    for col_index, row in enumerate(xlsx_file):
        for row_index, value in enumerate(row):
            cell = sheet.cell(row=row_index+1, column=col_index+1)
            cell.value = value
            
    sheet.delete_rows(3)
    sheet["A3"] = "User Id"
    
    create_add_row(3,2,id_list)
    
    create_add_row(4,2,user_pers)
    
    sheet.move_range("A1:F1", rows=+4,translate=False)
    sheet.move_range("B4:F4", rows=-3)
    sheet.move_range("A5:F5", rows=-1)



    
    
    # rows = sheet.max_row
    # cols = sheet.max_column
    # print(rows)
    # print(cols)
    # name_of_fields = []
    # fields_values = []
    # for i in range(1, rows+1):
    #     value_row = []
    #     for j in range(1, cols+1):
    #         cell = sheet.cell(row=j, column=i)
    #         value = str(cell.value)
    #         if i == 2:
    #             name_of_fields.append(value)
    #         else:
    #             value_row.append(value)
    #     if i != 2 and i!=1:
    #             fields_values.append(value_row)
    # print(fields_values)


    # i = 0
    # for row in fields_values:
    #     row[i]
    #     sheet.append(row)
    #     i+=1
    # i=0
    # for i, statN in enumerate(fields_values): 
    #     sheet.cell(row=i+1, column=1).value = statN 
    #     i+=1
   
   


# for row in sheet['A1:G37']:
#   for cell in row:
#     cell.value = None

       
# sheet["A2"] = 100
# sheet[1][0].value = "world"
# sheet.cell(row=1,column=1).value = "Hola"

book.save("data_name.xlsx")




# df = pd.read_excel('data_name.xlsx')
# df.T.to_excel('data_name.xlsx', index=False)
# book = openpyxl.load_workbook('data_name.xlsx')
# sheet = book.active
# sheet.move_range("B5:F5", rows=-4)
# sheet["A1"]=None
# book.save('data_name.xlsx')


