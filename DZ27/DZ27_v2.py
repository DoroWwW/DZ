import tkinter as tk
import openpyxl as xl
from tkinter import filedialog as fd
import os
import datetime
from dateutil.relativedelta import relativedelta

wb_obj, sheet_obj = None, None


def main_menu():
    while True:
        option = input(
            """Welcome to the XLSX reader 228-ULTRA
            1. Import file .xlsx
            2. Create new file .xlsx
            3. Exit\n"""
        )
        if option == "1":
            import_file()
        elif option == "2":
            create_file()
        elif option == "3":
            exit()
        else:
            clear_cmd()
            print("Choose the right option!")
            continue

        if wb_obj is not None and sheet_obj is not None:
            functional_menu()
        else:
            clear_cmd()
            print("File was not imported/created")


def functional_menu():
    while True:
        option = input(
            """Welcome to the XLSX reader 228-ULTRA
            1. Save
            2. Add
            3. Search
            4. Back
            5. Exit\n"""
        )

        if option == "1":
            save()
        elif option == "2":
            add()
        elif option == "3":
            search()
        elif option == "4":
            return
        elif option == "5":
            exit()

        clear_cmd()


def add():
    while True:
        name = input("Enter name: ")
        if not name.isalpha():
            print("Incorrect name")
            continue
        while True:
            surname = input("Enter surname (optional): ")
            if not surname.isalpha() and surname != "":
                print("Incorrect surname")
                continue
            break
        while True:
            second_name = input("Enter second name (optional): ")
            if not second_name.isalpha() and second_name != "":
                print("Incorrect second name")
                continue
            break
        while True:
            birth_date = input(
                "Enter birth date (DD.MM.YYYY or DD MM YYY or DD/MM/YYYY or dd-mm-YYYY): "
            )
            if birth_date is not None and birth_date != "":
                if birth_date.count(".") == 2:
                    birth_date = birth_date.split(".")
                elif birth_date.count(" ") == 2:
                    birth_date = birth_date.split(" ")
                elif birth_date.count("/") == 2:
                    birth_date = birth_date.split("/")
                elif birth_date.count("-") == 2:
                    birth_date = birth_date.split("/")
                try:
                    birth_day = int(birth_date[0])
                    birth_month = int(birth_date[1])
                    birth_year = int(birth_date[2])
                    break
                except Exception:
                    print("Not valid data")
        while True:
            gender = input("Enter gender (Male/Female or m/f): ")
            if gender.lower() == "male" or gender.lower() == "m":
                gender = "m"
                break
            elif gender.lower() == "female" or gender.lower() == "f":
                gender = "f"
                break
            else:
                print("Not valid gender")
        while True:
            death_date = input(
                'Enter death date (DD.MM.YYYY or DD MM YYY or DD/MM/YYYY or dd-mm-YYYY or print "no" if none): '
            )
            if death_date is not None and death_date != "":
                if death_date.lower() == "no" or death_date.lower() == "n":
                    death_date = None
                    break
                elif death_date.count(".") == 2:
                    death_date = death_date.split(".")
                elif death_date.count(" ") == 2:
                    death_date = death_date.split(" ")
                elif death_date.count("/") == 2:
                    death_date = death_date.split("/")
                elif death_date.count("-") == 2:
                    death_date = death_date.split("/")
                try:
                    death_day = int(death_date[0])
                    death_month = int(death_date[1])
                    death_year = int(death_date[2])
                    break
                except Exception:
                    print("Not valid data")
        sheet_obj.cell(row=sheet_obj.max_row+1, column=1).value = name
        sheet_obj.cell(row=sheet_obj.max_row, column=2).value = surname
        sheet_obj.cell(row=sheet_obj.max_row, column=3).value = second_name
        sheet_obj.cell(
            row=sheet_obj.max_row, column=4
        ).value = f"{birth_day}/{birth_month}/{birth_year}"
        if death_date is None:
            sheet_obj.cell(row=sheet_obj.max_row, column=5).value = ""
        else:
            sheet_obj.cell(
                row=sheet_obj.max_row, column=5
            ).value = f"{death_day}/{death_month}/{death_year}"
        sheet_obj.cell(row=sheet_obj.max_row, column=6).value = gender
        
        return


def search():
    while True:
        request = input("Enter name/surname/second name: ")
        if request is not None and request != "":
            request = request.lower()
            for row in sheet_obj.rows:
                if row[1].value is None:
                    row[1].value = ""
                if row[2].value is None:
                    row[2].value = ""

                if (
                    request in row[0].value.lower()
                    or request in row[1].value.lower()
                    or request in row[2].value.lower()
                ):
                    birth_date = row[3].value.split("/")
                    birth_day = birth_date[0]
                    birth_month = birth_date[1]
                    birth_year = birth_date[2]
                    start_date = datetime.date(
                        int(birth_year), int(birth_month), int(birth_day)
                    )

                    if row[4].value is None or row[4].value == "":
                        end_date = datetime.date.today()
                        years_lived = relativedelta(end_date, start_date).years
                        print(
                            f"\t{row[0].value} {row[1].value} {row[2].value} {row[5].value}, {row[3].value}, Years Lived: {years_lived}"
                        )
                    else:
                        death_date = row[4].value.split("/")
                        death_day = death_date[0]
                        death_month = death_date[1]
                        death_year = death_date[2]
                        end_date = datetime.date(
                            int(death_year), int(death_month), int(death_day)
                        )
                        years_lived = relativedelta(end_date, start_date).years
                        print(
                            f"\t{row[0].value} {row[1].value} {row[2].value} {row[5].value}, {row[3].value}, {row[4].value}, Years Lived: {years_lived}"
                        )
            input("Press any button")
            return
        else:
            print("User does not exist")
                        


def clear_cmd():
    os.system("cls")


def exit():
    os._exit(0)


def save():
    filename = input("Enter file name: ")
    wb_obj.save(f"{filename}.xlsx")


def create_file():
    global wb_obj, sheet_obj
    wb_obj = xl.Workbook()
    sheet_obj = wb_obj.active
    sheet_obj.title = "User Data"
    sheet_obj.cell(row=sheet_obj.max_row, column=1).value = "Name"
    sheet_obj.cell(row=sheet_obj.max_row, column=2).value = "Surname"
    sheet_obj.cell(row=sheet_obj.max_row, column=3).value = "Second Nam"
    sheet_obj.cell(row=sheet_obj.max_row, column=4).value = "Date of birth"
    sheet_obj.cell(row=sheet_obj.max_row, column=5).value = "Date of death"
    sheet_obj.cell(row=sheet_obj.max_row, column=6).value = "Gender"


def import_file():
    global wb_obj, sheet_obj
    while True:
        root = tk.Tk()
        root.overrideredirect(True)
        root.attributes("-alpha", 0)

        filetypes = (("XLSX Files", "*.xlsx"),)

        filepath = fd.askopenfilename(
            title="Open a file", initialdir="./", filetypes=filetypes
        )

        root.destroy()

        if filepath != "":
            wb_obj = xl.load_workbook(filepath)

            if "User Data" in wb_obj.sheetnames:
                sheet_obj = wb_obj["User Data"]
            else:
                sheet_obj = wb_obj.create_sheet("User Data")

            return
        return


main_menu()
