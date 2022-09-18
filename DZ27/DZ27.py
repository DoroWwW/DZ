
from ast import While
import subprocess
import sys

import openpyxl
import os



#дозапись 
def xlsx_writer(data_NS:list):
    
    # def create_add_row(row_number,your_column,a):
    for  row in data_NS:
        your_column = sheet_1.max_column
        for i, value in enumerate(row):
            sheet_1.cell(column=your_column, row=i).value = value 


def personality_loop(data:list):
    while True:
            name = input("Enter the name: ")
            if not name.isalpha:
                print("inc input!")
                continue
            surname = input("Enter the surname(optoinal): ")  
            if not surname.isalpha and not surname is None:
                print("inc input!!")
                continue
            f_name = input("Enter the F. name(optoinal): ")  
            if not f_name.isalpha and not f_name is None:
                print("inc input!!!")
            answer =input("Want to add another?(y/n): ")
            if answer.lower() in "n":
                person = Person(name, surname, f_name)
                
                return data.append(person.listmaker(name, surname, f_name))
                
            else:
                person = Person(name, surname, f_name)
                #person.listmaker(name, surname, f_name)
                data.append(person.listmaker(name, surname, f_name))
                continue


class Person:
    def __init__(self, name, surname=None, f_name=None):
        
        self.name = name
        self.surname = surname
        self.f_name = f_name
    def listmaker(self,name,surname,f_name):
        a = [name,surname,f_name]
        return a

while True:
    ent = input("Import data: 1\nCreate a new sheet: 2\nExit: 0\n")
    if not ent in ("1","2","0"):
        print("inc input")
        continue    
    if ent == "1":
        #open xlsx file
        pass
        while True:
            ent = input("Search: 1\nAdd new member: 2\n")
            if not ent in ("1","2"):
                print("inc input")
                continue
            if ent == "1":
                pass
            else:
                data_NS = []
                personality_loop(data_NS)
                #open n path for xlsx file
                xlsx_writer(data_NS)
        # os.system("start D:\GitHub reps\DZ")
        #-----------------------------------
        # path = r'D:\GitHub reps\DZ'
        # sys.path.append(path)
        # subprocess.Popen('explorer "D:\temp"')
        #-------------------------------------
    
    
    
    
    
    
    if ent == "2":
        data_NS = [["Name","Surname","F.Name"]]
        personality_loop(data_NS)
        book = openpyxl.Workbook()
        sheet_1 = book.create_sheet("User Data")
        for col_index, row in enumerate(data_NS):
            for row_index, value in enumerate(row):
                cell = sheet_1.cell(row=row_index+1, column=col_index+1)
                cell.value = value
        
                 
        file_name = input("Enter the file name: ") 
        book.save(f"{file_name}.xlsx")



    